# -*- coding: utf-8 -*-
import os
from datetime import datetime
import webview
import platform
import subprocess

class FolderExtractorAPI:
    def __init__(self, window):
        self.window = window

    def scan_folder(self, payload):
        include_sub = payload.get('include_subfolders', True)
        exclude_hidden = payload.get('exclude_hidden', True)
        extract_mode = payload.get('extract_mode', 'file')
        target_folder = payload.get('folder_path', None) # 기존 선택된 경로 확인
        
        MAX_FILES_LIMIT = 20000 

        # 전달받은 경로가 없으면 탐색기를 띄움
        if not target_folder:
            result = self.window.create_file_dialog(webview.FOLDER_DIALOG)
            if not result:
                return {"success": False, "message": "폴더 선택이 취소되었습니다."}
            target_folder = result[0]

        folder_path = target_folder
        file_list = []
        is_limited = False

        try:
            if include_sub:
                for root, dirs, files in os.walk(folder_path):
                    if is_limited: break
                    
                    if exclude_hidden:
                        dirs[:] = [d for d in dirs if not d.startswith('.')]

                    if extract_mode == 'folder':
                        if exclude_hidden and os.path.basename(root).startswith('.'):
                            continue
                        file_list.append(self._get_folder_info(root))
                        if len(file_list) >= MAX_FILES_LIMIT:
                            is_limited = True
                            break
                    else:
                        for file in files:
                            if exclude_hidden and file.startswith('.'): continue
                            if exclude_hidden and file.lower() in ['desktop.ini', 'thumbs.db']: continue
                            
                            file_list.append(self._get_file_info(root, file))
                            if len(file_list) >= MAX_FILES_LIMIT:
                                is_limited = True
                                break
            else:
                for item_name in os.listdir(folder_path):
                    if exclude_hidden and item_name.startswith('.'): continue
                    if exclude_hidden and item_name.lower() in ['desktop.ini', 'thumbs.db']: continue
                    
                    item_path = os.path.join(folder_path, item_name)
                    
                    if extract_mode == 'folder' and os.path.isdir(item_path):
                        file_list.append(self._get_folder_info(item_path))
                        if len(file_list) >= MAX_FILES_LIMIT:
                            is_limited = True
                            break
                    elif extract_mode == 'file' and os.path.isfile(item_path):
                        file_list.append(self._get_file_info(folder_path, item_name))
                        if len(file_list) >= MAX_FILES_LIMIT:
                            is_limited = True
                            break
                        
            return {
                "success": True, 
                "data": file_list, 
                "folder": folder_path,
                "is_limited": is_limited
            }
        except Exception as e:
            return {"success": False, "message": f"탐색 중 오류: {str(e)}"}

    def _get_file_info(self, directory, filename):
        path = os.path.join(directory, filename)
        size_b = 0
        try:
            stats = os.stat(path)
            size_b = stats.st_size
            if size_b < 1024: size_str = f"{size_b} B"
            elif size_b < 1024 * 1024: size_str = f"{size_b/1024:.1f} KB"
            elif size_b < 1024 * 1024 * 1024: size_str = f"{size_b/(1024*1024):.2f} MB"
            else: size_str = f"{size_b/(1024*1024*1024):.2f} GB"

            mod_time = datetime.fromtimestamp(stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
        except:
            size_str = "알 수 없음"
            mod_time = "알 수 없음"
            
        ext = os.path.splitext(filename)[1].lower()

        return {
            "name": filename,
            "ext": ext,
            "size": size_str,
            "size_bytes": size_b, # 통계 및 정렬을 위한 원본 바이트 추가
            "modified": mod_time,
            "path": directory
        }

    def _get_folder_info(self, folder_path):
        try:
            stats = os.stat(folder_path)
            mod_time = datetime.fromtimestamp(stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
        except:
            mod_time = "알 수 없음"
        
        return {
            "name": os.path.basename(folder_path) or folder_path,
            "ext": "폴더",
            "size": "-",
            "size_bytes": 0,
            "modified": mod_time,
            "path": os.path.dirname(folder_path)
        }

    def export_excel(self, payload):
        data_list = payload.get('data', [])
        if not data_list:
            return {"success": False, "message": "내보낼 데이터가 없습니다."}

        save_path = self.window.create_file_dialog(
            webview.SAVE_DIALOG,
            file_types=('Excel Files (*.xlsx)', 'All files (*.*)'),
            save_filename='목록_추출결과.xlsx'
        )

        if not save_path:
            return {"success": False, "message": "저장이 취소되었습니다."}

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "추출목록"

            headers = ["연번", "이름", "확장자/종류", "크기", "수정일자", "위치(폴더)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill("solid", fgColor="217346") 
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            thin_border = Border(left=Side(style='thin', color="E0E0E0"),
                               right=Side(style='thin', color="E0E0E0"),
                               top=Side(style='thin', color="E0E0E0"),
                               bottom=Side(style='thin', color="E0E0E0"))

            for row_idx, item in enumerate(data_list, 2):
                row_data = [
                    row_idx - 1,
                    item.get('name', ''),
                    item.get('ext', ''),
                    item.get('size', ''),
                    item.get('modified', ''),
                    item.get('path', '')
                ]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx in [1, 3, 4, 5]:
                        cell.alignment = Alignment(horizontal="center")

            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 60

            wb.save(save_path[0])
            wb.close()
            
            if platform.system() == 'Windows':
                os.startfile(save_path[0])
            elif platform.system() == 'Darwin':
                subprocess.call(('open', save_path[0]))
                
            return {"success": True, "message": "엑셀 저장이 완료되었습니다!"}
        except Exception as e:
            return {"success": False, "message": f"엑셀 저장 오류: {str(e)}"}