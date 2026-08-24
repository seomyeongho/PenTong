import os
import glob
from datetime import datetime
import json
import zipfile
import re
import html
import webview

class ExcelMergerAPI:
    def __init__(self, window, data_dir="."):
        self.window = window
        self.db_path = os.path.join(data_dir, "pentong_excel_merger.db")
        self._db_initialized = False

    def _init_db(self):
        """SQLite 데이터베이스 초기화"""
        if self._db_initialized:
            return
            
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id TEXT PRIMARY KEY,
                name TEXT,
                data_json TEXT,
                last_modified TEXT
            )
        ''')
        conn.commit()
        conn.close()
        self._db_initialized = True

    def _get_raw_defined_names(self, filepath):
        """(핵심) 엑셀 내부 XML에서 이름 정의를 100% 강제 추출 (동일 이름 덮어쓰기 방지)"""
        raw_names = []
        try:
            import zipfile, re, html
            with zipfile.ZipFile(filepath, 'r') as z:
                workbook_xml = z.read('xl/workbook.xml').decode('utf-8')
                pattern = r'<[^>]*definedName([^>]*)>(.*?)</[^>]*definedName>'
                for match in re.finditer(pattern, workbook_xml, re.DOTALL):
                    attrs = match.group(1)
                    val = html.unescape(match.group(2))
                    name_m = re.search(r'name="([^"]+)"', attrs)
                    local_id_m = re.search(r'localSheetId="(\d+)"', attrs)
                    if name_m:
                        d_name = html.unescape(name_m.group(1))
                        if not d_name.startswith('_xlnm'):
                            l_id = int(local_id_m.group(1)) if local_id_m else None
                            raw_names.append({"name": d_name, "value": val, "localSheetId": l_id})
        except Exception: pass
        return raw_names

    def _get_raw_tables(self, filepath):
        """(초핵심) openpyxl이 표(Table) 주소를 못 읽는 버그를 우회하여 XML에서 표 주소를 강제 직독직해!"""
        table_refs = {}
        try:
            import zipfile, re, html
            with zipfile.ZipFile(filepath, 'r') as z:
                for item in z.namelist():
                    if item.startswith('xl/tables/') and item.endswith('.xml'):
                        xml_str = z.read(item).decode('utf-8')
                        name_m = re.search(r'(?:name|displayName)="([^"]+)"', xml_str)
                        ref_m = re.search(r'ref="([^"]+)"', xml_str)
                        if name_m and ref_m:
                            t_name = html.unescape(name_m.group(1))
                            table_refs[t_name] = ref_m.group(1)
        except Exception: pass
        return table_refs

    # =========================================================================
    # [수정] 엑셀 수식, 구조적 참조(표), 일반 주소를 투명하게 번역하는 엔진
    # =========================================================================
    def _parse_excel_reference(self, wb, d_val, raw_tables_map=None):
        import re
        from openpyxl.utils.cell import range_boundaries, get_column_letter
        
        if raw_tables_map is None: raw_tables_map = {}
        if not d_val or not isinstance(d_val, str): return True, 150, f"문자열 아님", "", None

        d_val = d_val.strip()
        if d_val.startswith('='): d_val = d_val[1:]
            
        # 1. 표(Table) 구조적 참조 완벽 분석
        m = re.match(r'^([^!\[\]]+)\[([^\]]+)\]$', d_val)
        if m:
            t_name = m.group(1).strip()
            c_name = m.group(2).strip()
            for s_name in wb.sheetnames:
                if hasattr(wb[s_name], 'tables') and t_name in wb[s_name].tables:
                    tbl = wb[s_name].tables[t_name]
                    try:
                        # [핵심] openpyxl이 ref를 못 가져오면 XML 직독직해 값(raw_tables_map)으로 즉시 방어!
                        tbl_ref = getattr(tbl, 'ref', None) or raw_tables_map.get(t_name)
                        if not tbl_ref: return True, 150, f"표 '{t_name}' 범위 누락 (파싱 불가)", d_val, None
                            
                        min_c, min_r, max_c, max_r = range_boundaries(tbl_ref)
                        col_idx = -1
                        if hasattr(tbl, 'tableColumns') and tbl.tableColumns:
                            for i, col in enumerate(tbl.tableColumns):
                                if getattr(col, 'name', '') == c_name:
                                    col_idx = i
                                    break
                        if col_idx == -1:
                            cols = max_c - min_c + 1
                            return False, cols, f"표 '{t_name}' 열 인식 실패 (전체표 반환)", f"{s_name}!{tbl_ref}", s_name

                        tc = min_c + col_idx
                        dr = min_r + (1 if getattr(tbl, 'headerRowCount', 1) else 0)
                        return False, 1, f"표 '{t_name}'의 '{c_name}' 열", f"{get_column_letter(tc)}{dr}:{get_column_letter(tc)}{max_r}", s_name
                    except Exception as e:
                        return True, 150, f"표 참조 추적 실패: {str(e)}", d_val, None
            return True, 150, f"표 '{t_name}' 찾을 수 없음", d_val, None
            
        # 2. 일반 주소 (Sheet1!A1:B10)
        if '!' in d_val:
            s_name_raw, ref = d_val.rsplit('!', 1)
            s_name = s_name_raw.replace("'", "").replace('"', "")
            if '(' in ref or '[' in ref: return True, 150, "엑셀 수식/함수 포함 (동적 범위)", d_val, None
                
            ref = ref.replace('$', '').split(',')[0].strip()
            if ':' not in ref: ref = f"{ref}:{ref}"
            try:
                min_c, min_r, max_c, max_r = range_boundaries(ref)
                if max_c and min_c:
                    cols = max_c - min_c + 1
                    return False, cols, f"{get_column_letter(min_c)}{min_r} ~ {get_column_letter(max_c)}{max_r} ({cols}열)", ref, s_name
            except Exception as e: return True, 150, f"주소 계산 실패: {str(e)}", d_val, None
                
        # 3. 괄호가 포함된 기타 엑셀 함수 수식 (OFFSET 등)
        if '(' in d_val or '[' in d_val: return True, 150, "엑셀 수식/함수 포함 (동적 범위)", d_val, None
            
        # 4. 시트명이 없는 일반 주소 (A1:B10)
        ref = d_val.replace('$', '').split(',')[0].strip()
        if ':' not in ref: ref = f"{ref}:{ref}"
        try:
            min_c, min_r, max_c, max_r = range_boundaries(ref)
            if max_c and min_c:
                cols = max_c - min_c + 1
                return False, cols, f"{get_column_letter(min_c)}{min_r} ~ {get_column_letter(max_c)}{max_r} ({cols}열)", ref, None
        except Exception as e: return True, 150, f"단일 주소 에러: {str(e)}", d_val, None
            
        return True, 150, "예기치 않은 범위 구조", d_val, None

    def select_reference_file(self):
        import webview
        import os
        file_types = ('Excel Files (*.xlsx;*.xlsm)', 'All files (*.*)')
        result = self.window.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=file_types)
        
        if not result: return {"success": False, "message": "파일 선택이 취소되었습니다."}
            
        filepath = result[0]
        filename = os.path.basename(filepath)
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheets = wb.sheetnames
            
            named_ranges_by_sheet = {s: [] for s in sheets}
            global_named_ranges = [] 
            debug_names_info = [] 
            range_details = {} 
            
            # [핵심] XML 강제 표 파서 구동!
            raw_tables_map = self._get_raw_tables(filepath)
            
            for s in sheets:
                range_details[s] = {"maxCols": wb[s].max_column if wb[s].max_column else 50, "raw": f"시트 전체 ({s})", "isFormula": False, "parsedAddr": "시트 전체 데이터"}
            
            raw_names_list = self._get_raw_defined_names(filepath)
            for item in raw_names_list:
                d_name = item["name"]
                d_val = item["value"]
                l_id = item["localSheetId"]
                
                target_sheets = set()
                unique_key = d_name 
                
                if l_id is not None and 0 <= l_id < len(sheets):
                    target_sheets.add(sheets[l_id])
                    unique_key = f"{sheets[l_id]}!{d_name}" 
                else:
                    if d_name not in global_named_ranges: global_named_ranges.append(d_name)
                    if '!' in d_val:
                        parts = d_val.split(',')
                        for part in parts:
                            if '!' in part:
                                s_name = part.split('!')[0].replace('=', '').strip("'").strip('"')
                                if s_name in sheets: target_sheets.add(s_name)
                
                try:
                    is_formula, max_cols, parsed_addr, res_ref, res_sht = self._parse_excel_reference(wb, d_val, raw_tables_map)
                except Exception as e:
                    is_formula = True; max_cols = 150; parsed_addr = f"파이썬 에러: {str(e)}"
                
                range_details[unique_key] = {
                    "maxCols": max_cols if not is_formula else 150, "raw": str(d_val), "isFormula": is_formula, "parsedAddr": parsed_addr
                }
                
                for s in target_sheets:
                    if d_name not in named_ranges_by_sheet[s]: named_ranges_by_sheet[s].append(d_name)

            for sheet_name in sheets:
                ws = wb[sheet_name]
                if hasattr(ws, 'tables'):
                    for tbl_name, tbl_obj in ws.tables.items():
                        if tbl_name not in named_ranges_by_sheet[sheet_name]: named_ranges_by_sheet[sheet_name].append(tbl_name)
                        if tbl_name not in global_named_ranges: global_named_ranges.append(tbl_name)
                            
                        try:
                            # [핵심] openpyxl이 ref를 못 가져오면 XML 파서의 데이터로 즉시 대체!
                            ref_val = getattr(tbl_obj, 'ref', None) or raw_tables_map.get(tbl_name)
                            if not ref_val:
                                import html
                                ref_val = f"XML 파싱 불가 (Type: {html.escape(str(type(tbl_obj)))})"
                        except Exception as e:
                            ref_val = f"표 속성 접근 에러: {str(e)}"

                        raw_val = f"{sheet_name}!{ref_val}"
                        try:
                            is_formula, max_cols, parsed_addr, res_ref, res_sht = self._parse_excel_reference(wb, raw_val, raw_tables_map)
                        except Exception as e:
                            is_formula = True; max_cols = 150; parsed_addr = f"표 주소 분석 에러: {str(e)}"
                            
                        range_details[tbl_name] = {
                            "maxCols": max_cols if not is_formula else 150, "raw": raw_val, "isFormula": is_formula, "parsedAddr": parsed_addr
                        }

            wb.close()
            return {
                "success": True, "fileName": filename, "filepath": filepath, "sheets": sheets,
                "namedRangesBySheet": named_ranges_by_sheet, "globalNamedRanges": global_named_ranges,
                "rangeDetails": range_details, "debugNamesInfo": debug_names_info
            }
        except Exception as e:
            return {"success": False, "message": f"파일을 읽는 중 오류가 발생했습니다: {str(e)}"}

    def refresh_reference_file(self, payload):
        filepath = payload.get('filepath')
        import os
        if not filepath or not os.path.exists(filepath): return {"success": False, "message": "파일이 존재하지 않습니다."}
        filename = os.path.basename(filepath)
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheets = wb.sheetnames
            named_ranges_by_sheet = {s: [] for s in sheets}
            global_named_ranges = [] 
            range_details = {} 
            
            raw_tables_map = self._get_raw_tables(filepath)
            
            for s in sheets:
                range_details[s] = {"maxCols": wb[s].max_column if wb[s].max_column else 50, "raw": f"시트 전체 ({s})", "isFormula": False, "parsedAddr": "시트 전체 데이터"}
            
            raw_names_list = self._get_raw_defined_names(filepath)
            for item in raw_names_list:
                d_name = item["name"]
                d_val = item["value"]
                l_id = item["localSheetId"]
                
                target_sheets = set()
                unique_key = d_name
                
                if l_id is not None and 0 <= l_id < len(sheets):
                    target_sheets.add(sheets[l_id])
                    unique_key = f"{sheets[l_id]}!{d_name}" 
                else:
                    if d_name not in global_named_ranges: global_named_ranges.append(d_name)
                    if '!' in d_val:
                        parts = d_val.split(',')
                        for part in parts:
                            if '!' in part:
                                s_name = part.split('!')[0].replace('=', '').strip("'").strip('"')
                                if s_name in sheets: target_sheets.add(s_name)
                                    
                try:
                    is_formula, max_cols, parsed_addr, _, _ = self._parse_excel_reference(wb, d_val, raw_tables_map)
                except Exception as e:
                    is_formula = True; max_cols = 150; parsed_addr = f"파이썬 에러: {str(e)}"
                    
                range_details[unique_key] = {
                    "maxCols": max_cols if not is_formula else 150, "raw": str(d_val), "isFormula": is_formula, "parsedAddr": parsed_addr
                }
                                    
                for s in target_sheets:
                    if d_name not in named_ranges_by_sheet[s]: named_ranges_by_sheet[s].append(d_name)

            for sheet_name in sheets:
                ws = wb[sheet_name]
                if hasattr(ws, 'tables'):
                    for tbl_name, tbl_obj in ws.tables.items():
                        if tbl_name not in named_ranges_by_sheet[sheet_name]: named_ranges_by_sheet[sheet_name].append(tbl_name)
                        if tbl_name not in global_named_ranges: global_named_ranges.append(tbl_name)
                            
                        try:
                            ref_val = getattr(tbl_obj, 'ref', None) or raw_tables_map.get(tbl_name)
                            if not ref_val:
                                import html
                                ref_val = f"XML 파싱 불가 (Type: {html.escape(str(type(tbl_obj)))})"
                        except Exception as e:
                            ref_val = f"표 속성 접근 에러: {str(e)}"

                        raw_val = f"{sheet_name}!{ref_val}"
                        try:
                            is_formula, max_cols, parsed_addr, res_ref, res_sht = self._parse_excel_reference(wb, raw_val, raw_tables_map)
                        except Exception as e:
                            is_formula = True; max_cols = 150; parsed_addr = f"표 주소 분석 에러: {str(e)}"
                            
                        range_details[tbl_name] = {
                            "maxCols": max_cols if not is_formula else 150, "raw": raw_val, "isFormula": is_formula, "parsedAddr": parsed_addr
                        }

            wb.close()
            return {
                "success": True, "fileName": filename, "filepath": filepath, "sheets": sheets,
                "namedRangesBySheet": named_ranges_by_sheet, "globalNamedRanges": global_named_ranges, "rangeDetails": range_details 
            }
        except Exception as e:
            return {"success": False, "message": f"갱신 중 오류가 발생했습니다: {str(e)}"}

    def _scan_directory(self, folder_path):
        import os
        from datetime import datetime
        files_data = []
        try:
            for filename in os.listdir(folder_path):
                filepath = os.path.join(folder_path, filename)
                if not os.path.isfile(filepath): continue
                if filename.startswith("~$") or filename.startswith(".") or filename.lower() in ['desktop.ini', 'thumbs.db']: continue
                
                ext = os.path.splitext(filename)[1].lower()
                mtime = os.path.getmtime(filepath)
                dt = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
                
                has_error = False
                error_msg = ""
                supported_exts = ['.xlsx', '.xlsm', '.xltx', '.xltm']
                
                if ext not in supported_exts:
                    has_error = True
                    if ext in ['.xls', '.xlsb']: error_msg = f"미지원 구형/바이너리 엑셀({ext}): '다른 이름으로 저장(xlsx)' 필요"
                    else: error_msg = f"취합 불가 파일({ext}): 엑셀 파일이 아닙니다."
                        
                files_data.append({
                    "id": str(hash(filepath)), "name": filename, "path": filepath, "status": "pending", 
                    "lastModified": dt, "hasError": has_error, "errorMessage": error_msg
                })
            return {"success": True, "folder": folder_path, "files": files_data}
        except Exception as e: return {"success": False, "message": str(e)}

    def scan_folder(self):
        import webview
        result = self.window.create_file_dialog(webview.FOLDER_DIALOG)
        if not result: return {"success": False}
        return self._scan_directory(result[0])

    def refresh_folder(self, payload):
        folder_path = payload.get('folderPath')
        import os
        if not folder_path or not os.path.exists(folder_path): return {"success": False, "message": "폴더가 없습니다."}
        return self._scan_directory(folder_path)

    def save_project(self, project_data):
        self._init_db()
        try:
            import sqlite3
            import json
            from datetime import datetime
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            project_id = project_data.get('id', str(datetime.now().timestamp()))
            name = project_data.get('name', '제목 없음')
            data_json = json.dumps(project_data, default=str)
            last_modified = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('INSERT OR REPLACE INTO projects (id, name, data_json, last_modified) VALUES (?, ?, ?, ?)', (project_id, name, data_json, last_modified))
            conn.commit()
            conn.close()
            return {"success": True, "message": "저장 완료"}
        except Exception as e: return {"success": False, "message": str(e)}

    def get_projects(self):
        self._init_db()
        try:
            import sqlite3
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT id, name, data_json, last_modified FROM projects ORDER BY last_modified DESC')
            rows = cursor.fetchall()
            conn.close()
            projects = []
            for row in rows:
                try:
                    import json
                    proj = json.loads(row[2])
                    proj['id'] = row[0]
                    proj['name'] = row[1]
                    proj['last_modified'] = row[3]
                    projects.append(proj)
                except: pass
            return {"success": True, "projects": projects}
        except Exception as e: return {"success": False, "message": str(e)}

    def delete_project(self, payload):
        project_id = payload.get('id')
        self._init_db()
        import sqlite3
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM projects WHERE id=?', (project_id,))
            conn.commit()
            conn.close()
            return {"success": True}
        except Exception as e: return {"success": False, "message": str(e)}

    def rename_project(self, payload):
        project_id = payload.get('id')
        new_name = payload.get('name')
        self._init_db()
        import sqlite3
        import json
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT data_json FROM projects WHERE id=?', (project_id,))
            row = cursor.fetchone()
            if row:
                proj_data = json.loads(row[0])
                proj_data['name'] = new_name
                new_json = json.dumps(proj_data, default=str)
                cursor.execute('UPDATE projects SET name=?, data_json=? WHERE id=?', (new_name, new_json, project_id))
                conn.commit()
            conn.close()
            return {"success": True}
        except Exception as e: return {"success": False, "message": str(e)}

    def execute_merge(self, payload):
        import traceback
        import json
        import hashlib
        import os
        import time as pytime 
        from datetime import datetime, date, time
        
        last_update = [0]
        def _update_progress(percent, text, force=False):
            current_time = pytime.time()
            if force or (current_time - last_update[0]) > 0.1:
                try:
                    safe_text = json.dumps(text)
                    self.window.evaluate_js(f"if(window.updateMergeProgress) window.updateMergeProgress({percent}, {safe_text});")
                    last_update[0] = current_time
                except: pass
                
        try:
            import openpyxl
            def _safe_value(val):
                if val is None: return ""
                if isinstance(val, (datetime, date, time)): return str(val)
                return val
                
            # [핵심] 날짜 텍스트 처리 및 기호 해석 순서 버그가 완벽하게 수정된 조건 판별 엔진!
            def _check_condition(x, cond_expr):
                import re, fnmatch
                from datetime import datetime
                
                cond_expr = str(cond_expr).strip()
                # [수정] 정규식 기호 판별 순서 변경! (긴 기호를 먼저 판별하도록 하여 <> 가 < 와 > 로 쪼개지는 버그 완벽 해결)
                m = re.match(r'^(>=|<=|<>|!=|>|<|=)?(.*)$', cond_expr)
                if not m: return False
                op = m.group(1) or '=='
                if op == '=': op = '=='
                if op == '<>': op = '!='
                target = m.group(2).strip()
                
                # 사용자가 "해당 없음" 처럼 엑셀 습관대로 따옴표를 씌웠다면 쿨하게 벗겨줍니다!
                if (target.startswith('"') and target.endswith('"')) or (target.startswith("'") and target.endswith("'")):
                    target = target[1:-1]

                x_str = str(x) if x is not None else ""
                x_str_clean = x_str.strip()
                
                # 타겟이 비어있을 때 (예: <>, =)
                if target == "":
                    if op == '==': return x_str_clean == ""
                    if op == '!=': return x_str_clean != ""
                    return False

                if x_str_clean == "": return op == '!='

                def _to_float(val):
                    try: return float(val)
                    except:
                        v = str(val).strip()
                        # "2025-04-15" 같은 날짜 텍스트를 엑셀 시리얼 숫자로 강제 변환!
                        if len(v) >= 10 and v[4] == '-' and v[7] == '-':
                            try:
                                if len(v) == 10: dt = datetime.strptime(v, "%Y-%m-%d")
                                else: dt = datetime.strptime(v[:19], "%Y-%m-%d %H:%M:%S")
                                delta = dt - datetime(1899, 12, 30)
                                return float(delta.days) + (delta.seconds / 86400.0)
                            except: pass
                    raise ValueError

                try:
                    target_num = float(target)
                    try:
                        x_num = _to_float(x_str_clean)
                        if op == '==': return x_num == target_num
                        elif op == '!=': return x_num != target_num
                        elif op == '>': return x_num > target_num
                        elif op == '<': return x_num < target_num
                        elif op == '>=': return x_num >= target_num
                        elif op == '<=': return x_num <= target_num
                    except ValueError:
                        return op == '!='
                except ValueError:
                    if op == '==' and ('*' in target or '?' in target): return fnmatch.fnmatch(x_str_clean, target)
                    elif op == '!=' and ('*' in target or '?' in target): return not fnmatch.fnmatch(x_str_clean, target)
                    else:
                        if op == '==': return x_str_clean == target
                        elif op == '!=': return x_str_clean != target
                        elif op == '>': return x_str_clean > target
                        elif op == '<': return x_str_clean < target
                        elif op == '>=': return x_str_clean >= target
                        elif op == '<=': return x_str_clean <= target
                return False

            project_data = payload.get('project', {})
            files = payload.get('files', [])
            sheets_data = project_data.get('sheetsData', [])
            
            if not files: return {"success": False, "message": "취합할 파일이 없습니다."}
            selected_sheets = [s for s in sheets_data if s.get('selected')]
            if not selected_sheets: return {"success": False, "message": "취합할 시트가 선택되지 않았습니다."}

            merged_data = { s.get('targetName', s['name']): [] for s in selected_sheets }
            error_logs = []
            
            summary_config = project_data.get('summarySheet', {})
            summary_enabled = summary_config.get('enabled', False)
            summary_groups = summary_config.get('groups', [])
            if summary_enabled: merged_data["__summary__"] = []

            logs = project_data.get('logs', [])
            if not isinstance(logs, list): logs = []
            def add_log(msg, ltype="info"): logs.append({"time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "msg": msg, "type": ltype})

            # [핵심 픽스] 총괄표 설정(summary_config)이 1글자라도 변경되면 캐시를 무조건 폐기하고 재계산하도록 해시 맵 구조를 업그레이드!
            current_rules_dict = {
                "sheets": sheets_data,
                "summary": summary_config
            }
            current_rules_str = json.dumps(current_rules_dict, sort_keys=True)
            current_rules_hash = hashlib.md5(current_rules_str.encode('utf-8')).hexdigest()
            file_cache = project_data.get('fileCache', {})
            
            full_refresh_requested = payload.get('fullRefresh', False)
            if full_refresh_requested:
                add_log("사용자 요청: 모든 파일 캐시를 비우고 '전체 새로 취합'을 실행합니다.", "info")
                file_cache = {}
                project_data['rulesHash'] = current_rules_hash
            elif project_data.get('rulesHash') != current_rules_hash:
                add_log("취합 설정(디자인)이 변경되어 모든 파일을 새로 취합합니다.", "info")
                file_cache = {}
                project_data['rulesHash'] = current_rules_hash

            _update_progress(2, "대상 폴더 및 파일 리스트 분석 중...", force=True)

            fileindex_cols = {}
            for sheet_rule in selected_sheets:
                d_name = sheet_rule.get('targetName', sheet_rule['name'])
                fileindex_cols[d_name] = [i for i, col in enumerate(sheet_rule.get('layout', [])) if col.get('type') == 'meta' and col.get('cellType') == 'fileindex']

            for file_info in files:
                filepath = file_info.get('path')
                if filepath and os.path.exists(filepath):
                    try:
                        mtime = os.path.getmtime(filepath)
                        dt = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
                        file_info['lastModified'] = dt
                    except: pass

            file_sequence = 0
            new_file_cache = {}
            parsed_files = []
            unchanged_files = []
            total_files = len(files)
            
            for file_info in files:
                if file_info.get('excluded'): continue 
                
                file_info['hasError'] = False  
                file_info['hasWarning'] = False
                file_info['warningMessage'] = ""
                file_info['errorMessage'] = ""
                
                filepath = file_info.get('path')
                mtime = file_info.get('lastModified')
                filename = file_info.get('name')
                file_sequence += 1
                
                pct = 5 + int((file_sequence / total_files) * 90)
                _update_progress(pct, f"[{filename}] 취합 중... ({file_sequence}/{total_files}개)")
                
                if not os.path.exists(filepath):
                    error_logs.append(f"[{filename}] 파일을 찾을 수 없습니다.")
                    add_log(f"[{filename}] 파일을 찾을 수 없습니다.", "error")
                    file_info['hasError'] = True 
                    file_info['errorMessage'] = "파일을 찾을 수 없습니다."
                    file_info['mergeTime'] = None 
                    continue

                if filepath in file_cache and file_cache[filepath].get('lastModified') == mtime:
                    cached_entry = file_cache[filepath]
                    cached_data = cached_entry.get('data', {})
                    cached_warnings = cached_entry.get('warnings', "")
                    has_real_data = cached_entry.get('hasRealData')
                    cached_added_rows = 0  
                    
                    for d_name, rows in cached_data.items():
                        if d_name in merged_data:
                            if d_name == "__summary__":
                                merged_data[d_name].extend(rows) 
                            else:
                                idx_list = fileindex_cols.get(d_name, [])
                                for r in rows:
                                    row_copy = list(r)
                                    for idx in idx_list:
                                        if idx < len(row_copy): row_copy[idx] = str(file_sequence)
                                    merged_data[d_name].append(row_copy)
                                    cached_added_rows += 1
                                
                    new_file_cache[filepath] = cached_entry
                    # [추가] 캐시에서 시트 수 복구
                    file_info['sheetCount'] = cached_entry.get('sheetCount', 0)
                    
                    if has_real_data is None: has_real_data = (cached_added_rows > 0)
                    if not has_real_data:
                        err_msg = f"[{filename}] 파일에서 오류 발생:\n  - 취합 누락: 시트가 없거나 유효한 데이터가 없음"
                        error_logs.append(err_msg)
                        add_log(err_msg, "error")
                        file_info['hasError'] = True
                        file_info['errorMessage'] = "취합 누락: 시트가 없거나 유효한 데이터가 없음"
                        file_info['mergeTime'] = None
                    else:
                        unchanged_files.append(filename)
                        file_info['hasError'] = False
                        if cached_warnings:
                            file_info['hasWarning'] = True
                            file_info['warningMessage'] = cached_warnings
                        file_info['mergeTime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    continue
                
                parsed_files.append(filename)
                file_extracted_data = { s.get('targetName', s['name']): [] for s in selected_sheets }
                file_specific_errors = []
                
                try:
                    raw_names_list = self._get_raw_defined_names(filepath)
                    raw_tables_map = self._get_raw_tables(filepath)
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                    sheets = wb.sheetnames
                    
                    # [추가] 시트 수 계산하여 파일 정보에 저장
                    file_info['sheetCount'] = len(sheets)
                    
                    file_missing_sheets = []
                    empty_sheets = []
                    file_added_rows = 0
                    file_real_data_count = 0  
                    
                    if summary_enabled:
                        summary_row = []
                        if summary_config.get('includeFileIndex', False):
                            summary_row.append(str(file_sequence))
                        if summary_config.get('includeFileName', True):
                            summary_row.append(filename)
                            
                        for grp in summary_groups:
                            target_sheet = grp.get('targetSheet')
                            grp_range_str = grp.get('range', '')
                            grp_cols = grp.get('columns', [])
                            
                            grp_data = [] 
                            try:
                                ws_sum = wb[target_sheet] if target_sheet in sheets else wb[sheets[0]]
                                
                                if grp.get('type') == 'named' and grp_range_str:
                                    target_sheet_idx = sheets.index(target_sheet) if target_sheet in sheets else -1
                                    scoped_names = {item["name"]: item["value"] for item in raw_names_list if item["localSheetId"] in [None, target_sheet_idx]}
                                    if grp_range_str in scoped_names:
                                        d_val = scoped_names[grp_range_str]
                                        is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, d_val, raw_tables_map)
                                        if res_sht and res_sht in sheets: ws_sum = wb[res_sht]
                                        grp_range_str = res_ref
                                    elif hasattr(ws_sum, 'tables') and grp_range_str in ws_sum.tables:
                                        tbl_ref = getattr(ws_sum.tables[grp_range_str], 'ref', None) or raw_tables_map.get(grp_range_str)
                                        is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, f"{target_sheet}!{tbl_ref}", raw_tables_map)
                                        if res_sht and res_sht in sheets: ws_sum = wb[res_sht]
                                        grp_range_str = res_ref
                                        
                                elif grp.get('type') == 'manual' and grp_range_str:
                                    is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, grp_range_str, raw_tables_map)
                                    if res_sht and res_sht in sheets: ws_sum = wb[res_sht]
                                    grp_range_str = res_ref
                                    
                                if grp_range_str:
                                    if ':' not in grp_range_str: grp_range_str = f"{grp_range_str}:{grp_range_str}"
                                    for row in ws_sum[grp_range_str]:
                                        row_values = [_safe_value(cell.value) for cell in row]
                                        if not all(v == "" or str(v).strip() in ["0", "0.0"] for v in row_values):
                                            grp_data.append(row_values)
                            except Exception: pass
                                
                            for col_info in grp_cols:
                                c_idx = int(col_info.get('colIndex', 1)) - 1
                                func = col_info.get('func', 'first')
                                nth_row = int(col_info.get('nthRow', 1)) - 1 
                                cond_val = str(col_info.get('conditionVal', '')).strip()
                                cond_c_idx = int(col_info.get('condColIndex', 1)) - 1
                                
                                c_values = [r[c_idx] if c_idx < len(r) else "" for r in grp_data]
                                cond_values = [r[cond_c_idx] if cond_c_idx < len(r) else "" for r in grp_data]
                                
                                val_result = ""
                                if c_values:
                                    if func == 'first': val_result = c_values[0]
                                    elif func == 'nth': val_result = c_values[nth_row] if 0 <= nth_row < len(c_values) else ""
                                    elif func == 'counta': val_result = sum(1 for x in c_values if str(x).strip() != "")
                                    elif func in ['countif', 'sumif']:
                                        # [핵심 픽스] 드디어 파이썬에 '&' 기호를 쪼개서 다중 조건으로 해석하는 엔진 탑재!
                                        import re
                                        conds = []
                                        for part in cond_val.split('&'):
                                            part = part.strip()
                                            if not part: continue
                                            m_idx = re.match(r'^(\d+)\s*:(.*)$', part)
                                            if m_idx:
                                                conds.append((int(m_idx.group(1)) - 1, m_idx.group(2).strip()))
                                            else:
                                                # 열 번호가 생략된 단일 조건(예: >=100)이면 기본열로 매핑
                                                default_col = cond_c_idx if func == 'sumif' else c_idx
                                                conds.append((default_col, part))
                                                
                                        count = 0
                                        sum_val = 0
                                        for r in grp_data:
                                            match_all = True
                                            for c_i, c_expr in conds:
                                                x_val = r[c_i] if c_i < len(r) else ""
                                                if not _check_condition(x_val, c_expr):
                                                    match_all = False
                                                    break
                                            
                                            # 모든 조건을 통과(AND)한 경우에만 카운트 및 합산!
                                            if match_all and conds:
                                                count += 1
                                                try: sum_val += float(r[c_idx] if c_idx < len(r) else 0)
                                                except: pass
                                        
                                        val_result = sum_val if func == 'sumif' else count
                                    else:
                                        num_values = []
                                        for x in c_values:
                                            try: num_values.append(float(x))
                                            except: pass
                                        if func == 'count': val_result = len(num_values)
                                        elif func == 'sum': val_result = sum(num_values) if num_values else 0
                                        elif func == 'avg': val_result = sum(num_values) / len(num_values) if num_values else 0
                                        elif func == 'max': val_result = max(num_values) if num_values else ""
                                        elif func == 'min': val_result = min(num_values) if num_values else ""
                                summary_row.append(val_result)
                                
                        file_extracted_data["__summary__"] = [summary_row]
                        merged_data["__summary__"].append(summary_row)
                        file_real_data_count += 1 
                    
                    for sheet_rule in selected_sheets:
                        target_sheet_name = sheet_rule['name']
                        display_name = sheet_rule.get('targetName', target_sheet_name)
                        is_virtual = sheet_rule.get('isVirtual', False)
                        
                        if not is_virtual and target_sheet_name not in sheets:
                            file_missing_sheets.append(target_sheet_name) 
                            continue
                        
                        ws = wb[target_sheet_name] if target_sheet_name in sheets else wb[sheets[0]]
                        target_sheet_idx = sheets.index(target_sheet_name) if target_sheet_name in sheets else -1
                        
                        layout = sheet_rule.get('layout', [])
                        meta_values = {}
                        
                        for col in layout:
                            if col['type'] == 'meta':
                                cell_type = col.get('cellType', 'manual')
                                val = ""
                                if cell_type == 'filename': val = filename
                                elif cell_type == 'fileindex': val = str(file_sequence)
                                else:
                                    source_sheet_name = col.get('sheet') or target_sheet_name
                                    cell_addr_or_name = col.get('cell', '')
                                    if cell_addr_or_name:
                                        try:
                                            if cell_type == 'named':
                                                found_meta = False
                                                src_sheet_idx = sheets.index(source_sheet_name) if source_sheet_name in sheets else -1
                                                scoped_names = {}
                                                for item in raw_names_list:
                                                    if item["localSheetId"] is None: scoped_names[item["name"]] = item["value"]
                                                for item in raw_names_list:
                                                    if item["localSheetId"] == src_sheet_idx: scoped_names[item["name"]] = item["value"]
                                                        
                                                if cell_addr_or_name in scoped_names:
                                                    is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, scoped_names[cell_addr_or_name], raw_tables_map)
                                                    if res_sht and res_sht in wb.sheetnames:
                                                        c_addr = res_ref.split(':')[0]
                                                        val = _safe_value(wb[res_sht][c_addr].value)
                                                        found_meta = True
                                                        
                                                if not found_meta and source_sheet_name in wb.sheetnames and hasattr(wb[source_sheet_name], 'tables'):
                                                    if cell_addr_or_name in wb[source_sheet_name].tables:
                                                        t_ref = getattr(wb[source_sheet_name].tables[cell_addr_or_name], 'ref', None) or raw_tables_map.get(cell_addr_or_name)
                                                        if t_ref:
                                                            val = _safe_value(wb[source_sheet_name][t_ref.split(':')[0].replace('$', '')].value)
                                            else:
                                                is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, cell_addr_or_name, raw_tables_map)
                                                if res_sht and res_sht in wb.sheetnames:
                                                    val = _safe_value(wb[res_sht][res_ref.split(':')[0]].value)
                                                elif source_sheet_name in wb.sheetnames:
                                                    val = _safe_value(wb[source_sheet_name][res_ref.split(':')[0]].value)
                                        except Exception as inner_e:
                                            file_specific_errors.append(f"[{display_name} 시트] 부가항목 '{col.get('name')}' 추출 실패: {str(inner_e)}")
                                meta_values[col['id']] = val
                        
                        data_range_str = sheet_rule.get('range', '')
                        
                        if sheet_rule.get('type') == 'named' and data_range_str:
                            scoped_names = {item["name"]: item["value"] for item in raw_names_list if item["localSheetId"] in [None, target_sheet_idx]}
                            if data_range_str in scoped_names:
                                d_val = scoped_names[data_range_str]
                                is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, d_val, raw_tables_map)
                                if res_sht and res_sht in sheets: ws = wb[res_sht]
                                data_range_str = res_ref
                            elif hasattr(ws, 'tables') and data_range_str in ws.tables:
                                tbl_ref = getattr(ws.tables[data_range_str], 'ref', None) or raw_tables_map.get(data_range_str)
                                is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, f"{target_sheet_name}!{tbl_ref}", raw_tables_map)
                                if res_sht and res_sht in sheets: ws = wb[res_sht]
                                data_range_str = res_ref
                                
                        elif sheet_rule.get('type') == 'manual' and data_range_str:
                            is_form, m_cols, p_addr, res_ref, res_sht = self._parse_excel_reference(wb, data_range_str, raw_tables_map)
                            if res_sht and res_sht in sheets: ws = wb[res_sht]
                            data_range_str = res_ref
                            
                        if not data_range_str: continue

                        ignore_columns_str = sheet_rule.get('ignoreColumns', '')
                        ignore_indices = []
                        if ignore_columns_str:
                            for p in [x.strip() for x in ignore_columns_str.split(',')]:
                                if p.isdigit(): ignore_indices.append(int(p) - 1)
                        
                        sheet_row_count = 0
                        sheet_has_real_data = False
                        
                        try:
                            is_first_row_for_file = True
                            target_range = ws[data_range_str]
                            if not isinstance(target_range, (list, tuple)): target_range = [[target_range]]

                            for row_idx, row in enumerate(target_range):
                                row_values = [_safe_value(cell.value) for cell in row]
                                check_values = [v for i, v in enumerate(row_values) if i not in ignore_indices]
                                if not all(v == "" or str(v).strip() in ["0", "0.0"] for v in check_values):
                                    sheet_has_real_data = True
                                
                                if row_idx > 0:
                                    if all(v == "" or str(v).strip() in ["0", "0.0"] for v in check_values): continue

                                final_row = []
                                for col in layout:
                                    if col['type'] == 'meta':
                                        repeat_mode = col.get('repeatMode', 'all')
                                        if repeat_mode == 'first' and not is_first_row_for_file: final_row.append("")
                                        else: final_row.append(meta_values.get(col['id'], ''))
                                    elif col['type'] == 'data': final_row.extend(row_values)
                                file_extracted_data[display_name].append(final_row)
                                merged_data[display_name].append(final_row)
                                is_first_row_for_file = False
                                file_added_rows += 1 
                                sheet_row_count += 1
                        except Exception as inner_e:
                            file_specific_errors.append(f"[{display_name} 시트] 본문 데이터 추출 실패: {str(inner_e)}")
                            
                        if not is_virtual and target_sheet_name in sheets and not sheet_has_real_data: empty_sheets.append(target_sheet_name)
                        if sheet_has_real_data: file_real_data_count += 1
                            
                    wb.close()
                    
                    warnings_str = ""
                    if file_missing_sheets or empty_sheets:
                        w_list = []
                        if file_missing_sheets: w_list.append(f"시트 없음 ({', '.join(file_missing_sheets)})")
                        if empty_sheets: w_list.append(f"데이터 없음 ({', '.join(empty_sheets)})")
                        warnings_str = " / ".join(w_list)
                    
                    # [추가] 캐시에 시트 수도 함께 저장
                    new_file_cache[filepath] = {
                        'lastModified': mtime, 'data': file_extracted_data, 'warnings': warnings_str, 'hasRealData': file_real_data_count > 0,
                        'sheetCount': len(sheets)
                    }
                    
                    if file_real_data_count == 0 and not file_specific_errors:
                        if len(file_missing_sheets) == len(selected_sheets): reason = f"대상 시트가 엑셀 파일에 전혀 없음 ({', '.join(file_missing_sheets)})"
                        else: reason = "지정한 범위 내에 유효한 데이터가 없음 (또는 모두 빈칸)"
                        file_specific_errors.append(f"취합 누락 사유: {reason}")
                    elif warnings_str and file_real_data_count > 0:
                        file_info['hasWarning'] = True; file_info['warningMessage'] = warnings_str
                        add_log(f"[{filename}] 부분 누락/주의: {warnings_str}", "warning")
                        
                except Exception as file_e: file_specific_errors.append(f"파일 구조 읽기 오류: {str(file_e)}")

                if file_specific_errors:
                    file_info['hasError'] = True ; file_info['mergeTime'] = None 
                    err_detail = "\n  - ".join(file_specific_errors)
                    file_info['errorMessage'] = f"{err_detail}" 
                    grouped_err_msg = f"[{filename}] 파일에서 오류 발생:\n  - {err_detail}"
                    error_logs.append(grouped_err_msg)
                    add_log(grouped_err_msg, "error")
                else:
                    file_info['hasError'] = False; file_info['errorMessage'] = "" ; file_info['mergeTime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            _update_progress(95, "취합 결과 저장 중...", force=True)
            
            if parsed_files:
                if len(parsed_files) == 1: add_log(f"새로 취합됨: {parsed_files[0]}", "success")
                else: add_log(f"새로 취합됨: {parsed_files[0]} 외 {len(parsed_files)-1}개 파일", "success")

            project_data['fileCache'] = new_file_cache
            project_data['mergedData'] = merged_data
            project_data['files'] = files
            project_data['logs'] = logs[-1000:]
            
            save_result = self.save_project(project_data)
            if not save_result['success']: return save_result
                
            _update_progress(100, "취합 완료!", force=True)
            msg = f"취합 완료 (새로 취합: {len(parsed_files)}개, 캐시 사용: {len(unchanged_files)}개)"
            
            if error_logs:
                return {
                    "success": False, 
                    "message": f"작업이 완료되었으나, {len(error_logs)}개의 파일이 취합에서 누락되거나 오류가 발생했습니다.\n\n시트가 없거나 유효한 데이터가 없는 파일은 목록에 붉은색 아이콘으로 표시됩니다.\n\n어떤 파일이 무슨 이유로 누락되었는지 [로그 확인]을 눌러 점검해 주세요.", 
                    "project": project_data
                }
                
            return {"success": True, "message": msg, "project": project_data}
        except Exception as e:
            err_trace = traceback.format_exc()
            return {"success": False, "message": f"치명적 오류:\n{err_trace}"}

    def get_merged_data(self, payload):
        """[수정] DB에 저장된 취합 데이터를 화면으로 불러오기 (총괄표 맨 앞으로)"""
        project_id = payload.get('project', {}).get('id')
        if not project_id: return {"success": False, "message": "프로젝트 정보가 없습니다."}
        
        self._init_db()
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT data_json FROM projects WHERE id=?', (project_id,))
        row = cursor.fetchone()
        conn.close()
        
        if not row: return {"success": False, "message": "저장된 프로젝트를 찾을 수 없습니다."}
        
        try:
            import json
            proj_db = json.loads(row[0])
            merged_data = proj_db.get('mergedData', {})
            if not merged_data:
                return {"success": False, "message": "취합된 데이터가 없습니다. 먼저 [2. 대상 파일 관리] 탭에서 '취합 실행'을 진행해주세요."}
            
            summary_config = proj_db.get('summarySheet', {})
            sum_name = summary_config.get('sheetName', '총괄표')
            
            display_data = {}
            
            # [핵심] __summary__ (총괄표)가 있다면 가장 1순위로 dict에 집어넣어 화면의 1번 탭으로 만듦!
            if "__summary__" in merged_data:
                sheet_rows = []
                for r in merged_data["__summary__"]:
                    cleaned_row = [val.replace(" 00:00:00", "") if isinstance(val, str) and val.endswith(" 00:00:00") else val for val in r]
                    sheet_rows.append(cleaned_row)
                display_data[sum_name] = sheet_rows  # '__summary__' 대신 사용자가 지은 이름으로 변환
                
            for k, v in merged_data.items():
                if k == "__summary__": continue
                sheet_rows = []
                for r in v:
                    cleaned_row = [val.replace(" 00:00:00", "") if isinstance(val, str) and val.endswith(" 00:00:00") else val for val in r]
                    sheet_rows.append(cleaned_row)
                display_data[k] = sheet_rows

            return {"success": True, "data": display_data}
        except Exception as e:
            return {"success": False, "message": f"데이터 파싱 오류: {str(e)}"}

    def export_settings(self, payload):
        """[신규] 현재 취합 설정과 기준 양식을 묶어서 하나의 파일로 내보냅니다."""
        import zipfile, json, os, webview
        
        save_path = self.window.create_file_dialog(
            webview.SAVE_DIALOG,
            file_types=('Pentong Config Files (*.ptcfg)', 'ZIP Files (*.zip)', 'All files (*.*)'),
            save_filename='취합설정.ptcfg'
        )
        if not save_path: return {"success": False, "message": "취소됨"}

        project_data = payload.get('project', {})
        
        # [핵심] 데이터는 비우고, '원본 파일명'까지 확실하게 기록해서 저장!
        settings_to_save = {
            "name": project_data.get("name", "내보낸 취합 설정"),
            "sheetsData": project_data.get("sheetsData", []),
            "summarySheet": project_data.get("summarySheet", {}),
            "referenceFileName": project_data.get("referenceFileName", "")
        }

        out_zip = save_path[0]
        try:
            with zipfile.ZipFile(out_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr('settings.json', json.dumps(settings_to_save, ensure_ascii=False, indent=2))

                ref_file = project_data.get('referenceFile')
                template_path = ""
                if isinstance(ref_file, dict): template_path = ref_file.get('path', '')
                elif isinstance(ref_file, str): template_path = ref_file
                if not template_path: template_path = project_data.get('referenceFilePath', '')

                if template_path and os.path.exists(template_path):
                    zf.write(template_path, 'template.xlsx')
                    settings_to_save['has_template'] = True
                    zf.writestr('settings.json', json.dumps(settings_to_save, ensure_ascii=False, indent=2))
                    
            return {"success": True, "message": "설정과 양식이 성공적으로 내보내졌습니다!"}
        except Exception as e:
            return {"success": False, "message": f"내보내기 오류: {str(e)}"}

    def import_settings(self, payload):
        """[신규] 내보낸 설정 파일을 불러와 새로운 작업(프로젝트)으로 등록합니다."""
        import zipfile, json, os, uuid, time, webview
        from datetime import datetime
        
        open_path = self.window.create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=('Pentong Config Files (*.ptcfg)', 'ZIP Files (*.zip)', 'All files (*.*)')
        )
        if not open_path: return {"success": False, "message": "취소됨"}

        in_zip = open_path[0]
        try:
            with zipfile.ZipFile(in_zip, 'r') as zf:
                if 'settings.json' not in zf.namelist():
                    return {"success": False, "message": "유효한 취합 설정 파일이 아닙니다."}
                
                settings_data = json.loads(zf.read('settings.json').decode('utf-8'))
                
                app_data_dir = os.path.join(os.path.expanduser('~'), '.pentong_merger')
                os.makedirs(app_data_dir, exist_ok=True)
                
                template_path = ""
                if 'template.xlsx' in zf.namelist():
                    template_path = os.path.join(app_data_dir, f"template_{uuid.uuid4().hex[:8]}.xlsx")
                    with open(template_path, 'wb') as f:
                        f.write(zf.read('template.xlsx'))
                        
            # [핵심] 원본 파일명을 유지하고, 임시 폴더에서 풀렸다는 꼬리표(isImportedTemplate)를 붙임
            orig_name = settings_data.get("referenceFileName", "기준양식.xlsx")
            
            new_project = {
                "id": "proj_" + str(int(time.time() * 1000)),
                "name": settings_data.get("name", "가져온 설정") + " (불러옴)",
                "sheetsData": settings_data.get("sheetsData", []),
                "summarySheet": settings_data.get("summarySheet", {}),
                "files": [], "mergedData": {}, "fileCache": {},
                "referenceFile": {"path": template_path, "name": orig_name} if template_path else None,
                "referenceFileName": orig_name if template_path else "",
                "referenceFilePath": template_path,
                "isImportedTemplate": True if template_path else False
            }
            
            self._init_db()
            import sqlite3
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                'INSERT INTO projects (id, name, data_json, last_modified) VALUES (?, ?, ?, ?)',
                (new_project["id"], new_project["name"], json.dumps(new_project), datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
            conn.close()
            
            return {"success": True, "message": "설정을 성공적으로 불러와 새 작업 공간에 추가했습니다!"}
        except Exception as e:
            return {"success": False, "message": f"불러오기 오류: {str(e)}"}

    def export_reference_file(self, payload):
        """[신규] 숨겨진 폴더에 있는 기준 양식 파일을 사용자가 지정한 위치로 '다른 이름으로 저장' 합니다."""
        import webview, os, shutil
        project_data = payload.get('project', {})
        ref_path = project_data.get('referenceFilePath')
        ref_name = project_data.get('referenceFileName', '기준양식.xlsx')

        if not ref_path or not os.path.exists(ref_path):
            return {"success": False, "message": "저장할 기준 양식 원본 파일을 찾을 수 없습니다."}

        save_path = self.window.create_file_dialog(
            webview.SAVE_DIALOG,
            file_types=('Excel Files (*.xlsx;*.xlsm)', 'All files (*.*)'),
            save_filename=ref_name
        )

        if not save_path:
            return {"success": False, "message": "취소됨"}

        try:
            shutil.copy2(ref_path, save_path[0])
            return {"success": True, "message": "양식 파일이 성공적으로 저장되었습니다!"}
        except Exception as e:
            return {"success": False, "message": f"저장 중 오류 발생: {str(e)}"}

    def export_excel(self, payload):
        import webview
        import openpyxl
        import os
        import json
        import time
        from copy import copy
        from openpyxl.utils.cell import range_boundaries, get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        def _update_progress(percent, text):
            try:
                safe_text = json.dumps(text)
                self.window.evaluate_js(f"if(window.updateExportProgress) window.updateExportProgress({percent}, {safe_text});")
            except:
                pass

        def _copy_style(src_cell, tgt_cell):
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = copy(src_cell.number_format)
                tgt_cell.protection = copy(src_cell.protection)
                tgt_cell.alignment = copy(src_cell.alignment)

        def _get_col_width(ws, col_idx):
            letter = get_column_letter(col_idx)
            if letter in ws.column_dimensions:
                dim = ws.column_dimensions[letter]
                if dim.width is not None: return dim.width
            for dim in ws.column_dimensions.values():
                if getattr(dim, 'min', None) and getattr(dim, 'max', None):
                    if dim.min <= col_idx <= dim.max and dim.width is not None:
                        return dim.width
            return ws.sheet_format.defaultColWidth or 8.43

        def _get_row_height(ws, row_idx):
            if row_idx in ws.row_dimensions:
                dim = ws.row_dimensions[row_idx]
                if dim.height is not None: return dim.height
            return ws.sheet_format.defaultRowHeight or 15.0

        project_id = payload.get('project', {}).get('id')
        if not project_id: return {"success": False, "message": "프로젝트 정보가 없습니다."}
        
        self._init_db()
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT data_json FROM projects WHERE id=?', (project_id,))
        row = cursor.fetchone()
        conn.close()
        
        if not row: return {"success": False, "message": "저장된 프로젝트를 찾을 수 없습니다."}

        try:
            proj_db = json.loads(row[0])
            step = payload.get('step', 'export')

            if step == 'get_path':
                last_export_dir = proj_db.get('lastExportFolder', '')
                if not os.path.exists(last_export_dir):
                    last_export_dir = ''
                
                file_types = ('Excel Files (*.xlsx)', 'All files (*.*)')
                save_path = self.window.create_file_dialog(
                    webview.SAVE_DIALOG, 
                    directory=last_export_dir, 
                    file_types=file_types, 
                    save_filename='취합결과.xlsx'
                )
                if save_path:
                    return {"success": True, "path": save_path[0]}
                return {"success": False, "message": "저장이 취소되었습니다."}
            
            out_path = payload.get('savePath')
            if not out_path:
                return {"success": False, "message": "저장 경로가 전달되지 않았습니다."}

            _update_progress(5, "프로젝트 데이터 불러오는 중...")
            time.sleep(0.1) 

            sheets_data = proj_db.get('sheetsData', [])
            merged_data = proj_db.get('mergedData', {})
            files = proj_db.get('files', [])
            
            if not merged_data: return {"success": False, "message": "취합된 데이터가 없습니다. 먼저 '취합 실행'을 진행해주세요."}

            selected_sheets = [s for s in sheets_data if s.get('selected')]
            if not selected_sheets: return {"success": False, "message": "취합할 시트가 선택되지 않았습니다."}

            _update_progress(10, "기준 양식 템플릿 분석 중...")
            time.sleep(0.1)

            out_wb = openpyxl.Workbook()
            out_wb.remove(out_wb.active)
            
            # ========================================================
            # 총괄 시트(Summary) 내보내기 
            # ========================================================
            summary_config = proj_db.get('summarySheet', {})
            if summary_config.get('enabled', False) and "__summary__" in merged_data and merged_data["__summary__"]:
                sum_name = summary_config.get('sheetName', '총괄표')
                _update_progress(12, f"[{sum_name}] 총괄 시트 생성 중...")
                time.sleep(0.05)
                sum_ws = out_wb.create_sheet(title=sum_name, index=0)
                
                # 1행 헤더 디자인 및 작성 (사용자 옵션 반영)
                headers = []
                col_formats = [] # [신규] 셀 서식 리스트 수집
                
                if summary_config.get('includeFileIndex', False): 
                    headers.append('연번')
                    col_formats.append('')
                if summary_config.get('includeFileName', True): 
                    headers.append('파일명')
                    col_formats.append('')
                
                for grp in summary_config.get('groups', []):
                    for col in grp.get('columns', []): 
                        headers.append(col.get('title', '이름없음'))
                        col_formats.append(col.get('format', '').strip()) # 사용자가 지정한 셀 서식코드 담기
                        
                for c_idx, h_val in enumerate(headers, 1):
                    cell = sum_ws.cell(row=1, column=c_idx, value=h_val)
                    cell.fill = PatternFill("solid", fgColor="6B21A8") 
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # 열 너비 맞춤 설정
                    if h_val == '파일명': sum_ws.column_dimensions[get_column_letter(c_idx)].width = 28
                    elif h_val == '연번': sum_ws.column_dimensions[get_column_letter(c_idx)].width = 10
                    else: sum_ws.column_dimensions[get_column_letter(c_idx)].width = 14
                    
                sum_row_idx = 2
                
                meta_col_count = 0
                if summary_config.get('includeFileIndex', False): meta_col_count += 1
                if summary_config.get('includeFileName', True): meta_col_count += 1
                
                for r_data in merged_data["__summary__"]:
                    for c_idx, val in enumerate(r_data, 1):
                        cell = sum_ws.cell(row=sum_row_idx, column=c_idx, value=val)
                        bd = Side(style='thin', color="BFBFBF")
                        cell.border = Border(left=bd, right=bd, top=bd, bottom=bd)
                        cell.alignment = Alignment(vertical="center")
                        
                        fmt = col_formats[c_idx - 1] if (c_idx - 1) < len(col_formats) else ""
                        if fmt:
                            cell.number_format = fmt
                        # [수정] 서식을 지정하지 않으면 엑셀 기본값(General: 숫자는 숫자로, 문자는 문자로)을 그대로 유지합니다.
                    sum_row_idx += 1
            # ========================================================

            # [수정] 템플릿 경로 초기화 에러(UnboundLocalError) 완벽 방지
            template_path = ""
            ref_file_path = proj_db.get('referenceFilePath')
            ref_file = proj_db.get('referenceFile')
            
            if ref_file_path: 
                template_path = ref_file_path
            elif isinstance(ref_file, dict): 
                template_path = ref_file.get('path', '')
            elif isinstance(ref_file, str): 
                template_path = ref_file

            if not template_path or not os.path.exists(template_path):
                for f in files:
                    ext = os.path.splitext(f.get('name', ''))[1].lower()
                    if ext in ['.xlsx', '.xlsm', '.xltx', '.xltm'] and not f.get('hasError'):
                        template_path = f.get('path', '')
                        break
            
            template_wb = None


            template_raw_names = []
            if template_path and os.path.exists(template_path):
                try:
                    template_wb = openpyxl.load_workbook(template_path, data_only=True)
                    template_raw_names = self._get_raw_defined_names(template_path)
                except Exception:
                    pass

            total_sheets = len(selected_sheets)
            
            for i, sheet_rule in enumerate(selected_sheets):
                target_sheet_name = sheet_rule['name']
                display_name = sheet_rule.get('targetName', target_sheet_name)
                if display_name not in merged_data: continue
                
                base_pct = 15 + (i / total_sheets) * 75
                _update_progress(base_pct, f"[{display_name}] 시트 구조 생성 중...")
                time.sleep(0.05)
                
                out_ws = out_wb.create_sheet(title=display_name)
                layout = sheet_rule.get('layout', [])
                
                template_ws = None
                min_c = min_r = max_c = max_r = None
                
                if template_wb and target_sheet_name in template_wb.sheetnames:
                    template_ws = template_wb[target_sheet_name]
                    data_range_str = sheet_rule.get('range', '')
                    
                    if sheet_rule.get('type') == 'named' and data_range_str:
                        target_sheet_idx = template_wb.sheetnames.index(target_sheet_name)
                        scoped_names = {item["name"]: item["value"] for item in template_raw_names if item["localSheetId"] in [None, target_sheet_idx]}
                        if data_range_str in scoped_names:
                            data_range_str = scoped_names[data_range_str].split('!', 1)[-1].replace('=', '').replace('$', '')
                        elif hasattr(template_ws, 'tables') and data_range_str in template_ws.tables:
                            data_range_str = template_ws.tables[data_range_str].ref
                    
                    if data_range_str:
                        try:
                            if ':' not in data_range_str: data_range_str = f"{data_range_str}:{data_range_str}"
                            min_c, min_r, max_c, max_r = range_boundaries(data_range_str)
                        except: pass

                copy_min_r = max(1, min_r - 10) if min_r else 1
                copy_max_r = min_r - 1 if min_r and min_r > 1 else 1
                
                table_header_min_r = copy_max_r
                if template_ws and min_r:
                    for r in range(copy_max_r, copy_min_r - 1, -1):
                        has_content = False
                        for c in range(min_c, min_c + (max_c - min_c + 1)):
                            cell = template_ws.cell(row=r, column=c)
                            if cell.value is not None or (cell.has_style and (cell.border.bottom.style or cell.fill.patternType)):
                                has_content = True
                                break
                        if has_content: table_header_min_r = r
                        else: break

                sample_data_len = len(merged_data[display_name][0]) if merged_data[display_name] else 0
                meta_count = sum(1 for c in layout if c['type'] == 'meta')
                data_width = (max_c - min_c + 1) if (min_c and max_c) else 1
                if (sample_data_len - meta_count) > data_width: data_width = sample_data_len - meta_count

                current_out_col = 1
                col_mapping = {}
                data_out_start = None

                for col_item in layout:
                    out_letter = get_column_letter(current_out_col)
                    if col_item['type'] == 'meta':
                        out_ws.column_dimensions[out_letter].width = 16.0
                        if table_header_min_r < copy_max_r:
                            out_ws.merge_cells(start_row=table_header_min_r, start_column=current_out_col, end_row=copy_max_r, end_column=current_out_col)
                        
                        for r in range(table_header_min_r, copy_max_r + 1):
                            c_cell = out_ws.cell(row=r, column=current_out_col)
                            c_cell.fill = PatternFill("solid", fgColor="E2E8F0")
                            bd = Side(style='thin', color="A0AEC0")
                            c_cell.border = Border(left=bd, right=bd, top=bd, bottom=bd)
                            
                        top_cell = out_ws.cell(row=table_header_min_r, column=current_out_col)
                        top_cell.value = col_item.get('name', '부가항목')
                        top_cell.font = Font(bold=True, color="1E293B")
                        top_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        col_mapping[current_out_col] = None
                        current_out_col += 1
                    else:
                        if data_out_start is None: data_out_start = current_out_col
                        for dc in range(data_width):
                            src_c_idx = min_c + dc if min_c else 1
                            out_letter_data = get_column_letter(current_out_col)
                            if template_ws:
                                out_ws.column_dimensions[out_letter_data].width = _get_col_width(template_ws, src_c_idx)
                            col_mapping[current_out_col] = src_c_idx
                            current_out_col += 1

                if template_ws and min_r and min_c and data_out_start:
                    for r in range(copy_min_r, copy_max_r + 1):
                        out_ws.row_dimensions[r].height = _get_row_height(template_ws, r)
                        for c_offset in range(data_width):
                            src_cell = template_ws.cell(row=r, column=min_c + c_offset)
                            tgt_cell = out_ws.cell(row=r, column=data_out_start + c_offset)
                            tgt_cell.value = src_cell.value
                            _copy_style(src_cell, tgt_cell)
                    
                    for merged_range in template_ws.merged_cells.ranges:
                        m_min_c, m_min_r_m, m_max_c, m_max_r_m = range_boundaries(str(merged_range))
                        if m_min_r_m >= copy_min_r and m_max_r_m <= copy_max_r and m_min_c >= min_c and m_max_c <= max_c:
                            offset = data_out_start - min_c
                            try:
                                out_ws.merge_cells(
                                    start_row=m_min_r_m, start_column=m_min_c + offset,
                                    end_row=m_max_r_m, end_column=m_max_c + offset
                                )
                            except: pass

                out_row_idx = copy_max_r + 1
                total_rows = len(merged_data[display_name])
                
                for row_idx, row_data in enumerate(merged_data[display_name]):
                    if row_idx % 200 == 0:  
                        sheet_pct = ((row_idx / total_rows) if total_rows else 1) * (75 / total_sheets)
                        _update_progress(base_pct + sheet_pct, f"[{display_name}] 데이터 기록 중... ({row_idx}/{total_rows}행)")
                        
                    if template_ws and min_r: out_ws.row_dimensions[out_row_idx].height = _get_row_height(template_ws, min_r)
                    for c_idx, val in enumerate(row_data):
                        out_col_idx = c_idx + 1
                        tgt_cell = out_ws.cell(row=out_row_idx, column=out_col_idx)
                        if isinstance(val, str):
                            if val.endswith(" 00:00:00"):
                                try: val = datetime.strptime(val, "%Y-%m-%d 00:00:00").date()
                                except: pass
                            elif len(val) == 19 and val[4] == '-' and val[13] == ':':
                                try: val = datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
                                except: pass
                        tgt_cell.value = val

                        src_col_idx = col_mapping.get(out_col_idx)
                        if src_col_idx and template_ws and min_r:
                            src_cell = template_ws.cell(row=min_r, column=src_col_idx)
                            _copy_style(src_cell, tgt_cell)
                        else:
                            bd = Side(style='thin', color="BFBFBF")
                            tgt_cell.border = Border(left=bd, right=bd, top=bd, bottom=bd)
                            tgt_cell.alignment = Alignment(vertical="center")
                    out_row_idx += 1
                    
            _update_progress(95, "파일 저장 중 (잠시만 기다려주세요)...")
            time.sleep(0.1)
            
            if template_wb: template_wb.close()
            out_wb.save(out_path)
            
            proj_db['lastExportFolder'] = os.path.dirname(out_path)
            self.save_project(proj_db)
            
            _update_progress(100, "내보내기 완료!")
            return {"success": True, "message": "성공적으로 제목과 양식을 보존하여 내보냈습니다!"}
        except Exception as e:
            import traceback
            return {"success": False, "message": f"파일 저장 중 오류 발생: {str(e)}\n{traceback.format_exc()}"}